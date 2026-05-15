"""
PROMPT 14 — Personalized Guest Experience
- Token-based personalized invitation links: /invite/{slug}?g={token}
- Manual guest entry + CSV/Excel bulk import
- Per-guest voice message (recorded via browser MediaRecorder, stored as .webm)
- WhatsApp share helpers
"""
from __future__ import annotations

import io
import os
import csv
import uuid
import string
import secrets
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Optional, Dict, Any, Literal

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel, Field

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


VOICE_DIR = Path("/app/uploads/voice_messages")
VOICE_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
TOKEN_ALPHABET = string.ascii_letters + string.digits


def _new_token() -> str:
    return "".join(secrets.choice(TOKEN_ALPHABET) for _ in range(12))


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


class Guest(BaseModel):
    id: str = Field(default_factory=lambda: uuid.uuid4().hex)
    profile_id: str
    token: str = Field(default_factory=_new_token)
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    relationship: Optional[str] = None       # e.g., "Bride's cousin", "Family friend"
    table_number: Optional[str] = None
    seat_number: Optional[str] = None
    meal_preference: Literal["veg", "non_veg", "jain", "unspecified"] = "unspecified"
    events_invited: List[str] = []           # e.g., ["mehendi", "sangeet", "wedding", "reception"]
    voice_message_url: Optional[str] = None
    is_vip: bool = False
    invitation_sent: bool = False
    invitation_opened: bool = False
    rsvp_status: Optional[str] = None        # mirrored from RSVP collection if exists
    created_at: str = Field(default_factory=_now_iso)
    updated_at: str = Field(default_factory=_now_iso)


class GuestCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    relationship: Optional[str] = None
    table_number: Optional[str] = None
    seat_number: Optional[str] = None
    meal_preference: Literal["veg", "non_veg", "jain", "unspecified"] = "unspecified"
    events_invited: List[str] = []
    is_vip: bool = False


class GuestUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    relationship: Optional[str] = None
    table_number: Optional[str] = None
    seat_number: Optional[str] = None
    meal_preference: Optional[Literal["veg", "non_veg", "jain", "unspecified"]] = None
    events_invited: Optional[List[str]] = None
    is_vip: Optional[bool] = None
    invitation_sent: Optional[bool] = None


def _strip(d: dict) -> dict:
    d.pop("_id", None)
    return d


# ---------------------------------------------------------------------------
# Router builder (called from server.py)
# ---------------------------------------------------------------------------
def build_guest_router(db, require_admin, get_optional_admin=None) -> APIRouter:
    router = APIRouter(tags=["guests"])

    # ---- Helpers -----------------------------------------------------------
    async def _get_profile_owned(profile_id: str, admin_data: dict) -> dict:
        p = await db.profiles.find_one({"$or": [{"id": profile_id}, {"slug": profile_id}]}, {"_id": 0})
        if not p:
            raise HTTPException(404, "Profile not found")
        admin_id = admin_data.get("admin_id") or admin_data.get("id")
        if admin_data.get("role") not in ("super_admin", "superadmin") and p.get("admin_id") and p.get("admin_id") != admin_id:
            raise HTTPException(403, "Not your wedding")
        return p

    async def _get_profile_by_slug(slug: str) -> dict:
        p = await db.profiles.find_one({"$or": [{"id": slug}, {"slug": slug}]}, {"_id": 0})
        if not p:
            raise HTTPException(404, "Wedding not found")
        return p

    # ---- ADMIN ENDPOINTS ---------------------------------------------------
    @router.get("/admin/profiles/{profile_id}/guests")
    async def list_guests(profile_id: str, admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        cursor = db.guests.find({"profile_id": profile["id"]}, {"_id": 0}).sort("created_at", -1)
        guests = await cursor.to_list(2000)
        # Stats
        total = len(guests)
        vip_count = sum(1 for g in guests if g.get("is_vip"))
        sent_count = sum(1 for g in guests if g.get("invitation_sent"))
        opened_count = sum(1 for g in guests if g.get("invitation_opened"))
        with_voice = sum(1 for g in guests if g.get("voice_message_url"))
        return {
            "total": total,
            "vip": vip_count,
            "sent": sent_count,
            "opened": opened_count,
            "with_voice": with_voice,
            "guests": guests,
        }

    @router.post("/admin/profiles/{profile_id}/guests")
    async def create_guest(profile_id: str, payload: GuestCreate, admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        guest = Guest(profile_id=profile["id"], **payload.model_dump())
        doc = guest.model_dump()
        await db.guests.insert_one(doc)
        return _strip(doc)

    @router.put("/admin/profiles/{profile_id}/guests/{guest_id}")
    async def update_guest(profile_id: str, guest_id: str, payload: GuestUpdate,
                           admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        update = {k: v for k, v in payload.model_dump().items() if v is not None}
        update["updated_at"] = _now_iso()
        res = await db.guests.update_one(
            {"id": guest_id, "profile_id": profile["id"]},
            {"$set": update},
        )
        if res.matched_count == 0:
            raise HTTPException(404, "Guest not found")
        guest = await db.guests.find_one({"id": guest_id}, {"_id": 0})
        return guest

    @router.delete("/admin/profiles/{profile_id}/guests/{guest_id}")
    async def delete_guest(profile_id: str, guest_id: str, admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        # Also remove voice message file if present
        guest = await db.guests.find_one({"id": guest_id, "profile_id": profile["id"]}, {"_id": 0})
        if guest and guest.get("voice_message_url"):
            try:
                voice_path = VOICE_DIR / f"{guest['token']}.webm"
                if voice_path.exists():
                    voice_path.unlink()
            except Exception:
                pass
        res = await db.guests.delete_one({"id": guest_id, "profile_id": profile["id"]})
        if res.deleted_count == 0:
            raise HTTPException(404, "Guest not found")
        return {"success": True}

    @router.post("/admin/profiles/{profile_id}/guests/bulk-delete")
    async def bulk_delete_guests(profile_id: str, payload: Dict[str, Any],
                                  admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        ids = payload.get("guest_ids") or []
        res = await db.guests.delete_many({"profile_id": profile["id"], "id": {"$in": ids}})
        return {"deleted": res.deleted_count}

    # ---- Excel / CSV Template Download ------------------------------------
    @router.get("/admin/profiles/{profile_id}/guests/template")
    async def download_template(profile_id: str, format: str = "xlsx",
                                 admin_data: dict = Depends(require_admin)):
        await _get_profile_owned(profile_id, admin_data)
        headers = ["Name", "Phone", "Email", "Relationship", "Table", "Seat",
                   "Meal (veg/non_veg/jain)", "Events (comma-separated)", "VIP (yes/no)"]
        sample = ["Anaya Sharma", "+91 98765 43210", "anaya@example.com",
                  "Bride's cousin", "5", "12", "veg", "mehendi,sangeet,wedding,reception", "yes"]

        if format == "csv" or not _HAS_OPENPYXL:
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            writer.writerow(sample)
            return StreamingResponse(
                io.BytesIO(buf.getvalue().encode("utf-8")),
                media_type="text/csv",
                headers={"Content-Disposition": 'attachment; filename="guest_list_template.csv"'},
            )

        # XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Guests"
        ws.append(headers)
        ws.append(sample)
        # Column widths
        for i, w in enumerate([22, 18, 26, 22, 8, 8, 22, 36, 12], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        # Bold header
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="guest_list_template.xlsx"'},
        )

    # ---- CSV / XLSX Import ------------------------------------------------
    @router.post("/admin/profiles/{profile_id}/guests/import")
    async def import_guests(profile_id: str, file: UploadFile = File(...),
                             admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        filename = (file.filename or "").lower()
        raw = await file.read()
        rows: List[List[str]] = []
        try:
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                if not _HAS_OPENPYXL:
                    raise HTTPException(400, "Excel parser unavailable on server; please use CSV")
                wb = load_workbook(io.BytesIO(raw), data_only=True)
                ws = wb.active
                for r in ws.iter_rows(values_only=True):
                    rows.append([("" if v is None else str(v)).strip() for v in r])
            else:
                text = raw.decode("utf-8-sig", errors="replace")
                reader = csv.reader(io.StringIO(text))
                for r in reader:
                    rows.append([(c or "").strip() for c in r])
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(400, f"Could not parse file: {str(e)[:140]}")

        if not rows or len(rows) < 2:
            raise HTTPException(400, "File appears empty")

        # Detect header
        header_keys = [h.lower() for h in rows[0]]

        def col(row: List[str], names: List[str]) -> str:
            for n in names:
                if n in header_keys:
                    idx = header_keys.index(n)
                    if idx < len(row):
                        return row[idx]
            return ""

        added, skipped = [], []
        for r in rows[1:]:
            if not any(r):
                continue
            name = col(r, ["name", "full name"])
            phone = col(r, ["phone", "mobile", "phone number"])
            if not name or not phone:
                skipped.append({"row": r, "reason": "Missing required name or phone"})
                continue
            meal = (col(r, ["meal (veg/non_veg/jain)", "meal", "meal preference"]) or "unspecified").lower()
            if meal in ("non veg", "non-veg", "nonveg"):
                meal = "non_veg"
            if meal not in ("veg", "non_veg", "jain", "unspecified"):
                meal = "unspecified"
            events_raw = col(r, ["events (comma-separated)", "events", "events invited"])
            events = [e.strip().lower() for e in events_raw.split(",") if e.strip()]
            vip_raw = (col(r, ["vip (yes/no)", "vip"]) or "").lower().strip()
            is_vip = vip_raw in ("yes", "true", "1", "y")
            guest = Guest(
                profile_id=profile["id"],
                name=name,
                phone=phone,
                email=col(r, ["email"]) or None,
                relationship=col(r, ["relationship"]) or None,
                table_number=col(r, ["table"]) or None,
                seat_number=col(r, ["seat"]) or None,
                meal_preference=meal,  # type: ignore
                events_invited=events,
                is_vip=is_vip,
            )
            doc = guest.model_dump()
            await db.guests.insert_one(doc)
            added.append(_strip(doc))
        return {"added": len(added), "skipped": len(skipped), "guests": added, "skipped_rows": skipped}

    # ---- Voice Message Upload --------------------------------------------
    @router.post("/admin/profiles/{profile_id}/guests/{guest_id}/voice")
    async def upload_voice(profile_id: str, guest_id: str,
                            file: UploadFile = File(...),
                            admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        guest = await db.guests.find_one({"id": guest_id, "profile_id": profile["id"]}, {"_id": 0})
        if not guest:
            raise HTTPException(404, "Guest not found")
        raw = await file.read()
        if len(raw) > 5 * 1024 * 1024:  # 5 MB safety cap
            raise HTTPException(413, "Voice message too large (max 5 MB)")
        # Save under voice_messages/{wedding_id}/{guest_token}.webm
        wedding_dir = VOICE_DIR / profile["id"]
        wedding_dir.mkdir(parents=True, exist_ok=True)
        out_path = wedding_dir / f"{guest['token']}.webm"
        out_path.write_bytes(raw)
        url = f"/api/uploads/voice/{profile['id']}/{guest['token']}.webm"
        await db.guests.update_one(
            {"id": guest_id, "profile_id": profile["id"]},
            {"$set": {"voice_message_url": url, "updated_at": _now_iso()}},
        )
        return {"voice_message_url": url}

    @router.delete("/admin/profiles/{profile_id}/guests/{guest_id}/voice")
    async def delete_voice(profile_id: str, guest_id: str,
                            admin_data: dict = Depends(require_admin)):
        profile = await _get_profile_owned(profile_id, admin_data)
        guest = await db.guests.find_one({"id": guest_id, "profile_id": profile["id"]}, {"_id": 0})
        if not guest:
            raise HTTPException(404, "Guest not found")
        try:
            (VOICE_DIR / profile["id"] / f"{guest['token']}.webm").unlink(missing_ok=True)
        except Exception:
            pass
        await db.guests.update_one(
            {"id": guest_id, "profile_id": profile["id"]},
            {"$set": {"voice_message_url": None, "updated_at": _now_iso()}},
        )
        return {"success": True}

    @router.get("/uploads/voice/{profile_id}/{filename}")
    async def serve_voice(profile_id: str, filename: str):
        path = VOICE_DIR / profile_id / filename
        if not path.exists():
            raise HTTPException(404, "Voice not found")
        return FileResponse(path, media_type="audio/webm")

    # ---- WhatsApp Bulk-Helper --------------------------------------------
    @router.post("/admin/profiles/{profile_id}/guests/whatsapp-links")
    async def whatsapp_links(profile_id: str, payload: Dict[str, Any],
                              admin_data: dict = Depends(require_admin),
                              request: Request = None):
        profile = await _get_profile_owned(profile_id, admin_data)
        ids = payload.get("guest_ids") or []
        message_template = (payload.get("message") or
                            "You are warmly invited to {couple}'s wedding! Open your personal invitation here: {link}")
        couple = f"{profile.get('groom_name','')} & {profile.get('bride_name','')}".strip(" &")
        slug = profile.get("slug") or profile["id"]
        # Use frontend URL from env or fallback
        frontend_url = os.environ.get("FRONTEND_URL", "")
        # Pull from request host if not configured
        if not frontend_url and request is not None:
            frontend_url = str(request.base_url).rstrip("/")

        cursor = db.guests.find(
            {"profile_id": profile["id"], "id": {"$in": ids}} if ids else {"profile_id": profile["id"]},
            {"_id": 0},
        )
        guests = await cursor.to_list(2000)
        links = []
        for g in guests:
            link = f"{frontend_url}/invite/{slug}?g={g['token']}"
            msg = message_template.replace("{couple}", couple).replace("{link}", link).replace("{name}", g.get("name", ""))
            phone = (g.get("phone") or "").replace(" ", "").replace("-", "")
            wa = ""
            if phone:
                wa_phone = phone if phone.startswith("+") else f"+91{phone}" if not phone.startswith("91") else f"+{phone}"
                wa = f"https://wa.me/{wa_phone.replace('+', '')}?text={_url_encode(msg)}"
            links.append({
                "guest_id": g["id"], "name": g.get("name"), "phone": phone,
                "personal_link": link, "whatsapp_link": wa, "message": msg,
            })
        return {"links": links}

    # ---- PUBLIC: token-based personalized guest view ----------------------
    @router.get("/invite/{slug}/guest/{token}")
    async def get_personalized_guest(slug: str, token: str):
        profile = await _get_profile_by_slug(slug)
        guest = await db.guests.find_one(
            {"profile_id": profile["id"], "token": token},
            {"_id": 0, "email": 0, "phone": 0},  # don't leak PII
        )
        if not guest:
            raise HTTPException(404, "This invitation link is not valid")
        # Mark as opened (don't update if already opened)
        if not guest.get("invitation_opened"):
            await db.guests.update_one(
                {"id": guest["id"]},
                {"$set": {"invitation_opened": True, "opened_at": _now_iso()}},
            )
        return {
            "name": guest.get("name"),
            "relationship": guest.get("relationship"),
            "table_number": guest.get("table_number"),
            "seat_number": guest.get("seat_number"),
            "meal_preference": guest.get("meal_preference"),
            "events_invited": guest.get("events_invited", []),
            "voice_message_url": guest.get("voice_message_url"),
            "is_vip": bool(guest.get("is_vip")),
            "couple": f"{profile.get('groom_name','')} & {profile.get('bride_name','')}".strip(" &"),
        }

    return router


def _url_encode(s: str) -> str:
    import urllib.parse
    return urllib.parse.quote(s, safe="")
